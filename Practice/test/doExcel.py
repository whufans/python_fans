#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/10 10:13
# @Author  : fans
# @File    : doExcel.py
# @Software: PyCharm Community Edition
from openpyxl import load_workbook

class DoExcel:
	def __init__(self,filePath,sheetName):
		self.filePath=filePath
		self.sheetName=sheetName
	def doExcel(self):
		wb=load_workbook(self.filePath)
		sheet=wb[self.sheetName]
		testData=[]
		for i in range(2,sheet.max_row+1):
			subData={}
			subData['id'] = sheet.cell(i, 1).value
			subData['title'] = sheet.cell(i, 2).value
			subData['http_method'] = sheet.cell(i, 3).value
			subData['url'] = sheet.cell(i, 4).value
			subData['data'] = sheet.cell(i, 5).value
			subData['expected'] = sheet.cell(i, 6).value
			testData.append(subData)
		return testData
	def writeBack(self,row,column,value):
		wb = load_workbook(self.filePath)
		sheet = wb[self.sheetName]
		sheet.cell(row,column).value=value
		wb.save(self.filePath)

