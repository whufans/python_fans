#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/3 16:58
# @Author  : fans
# @File    : do_excel.py
# @Software: PyCharm Community Edition
from openpyxl import  load_workbook
class DoExcel:
	def __init__(self,filePath,sheetName):
		self.filePath=filePath
		self.sheetName=sheetName
	def do_excel(self):
		wb=load_workbook(self.filePath)
		sheet=wb[self.sheetName]
		test_data=[]
		for i in range(2,sheet.max_row+1):
			sub_data={}
			sub_data['id']=sheet.cell(i,1).value
			sub_data['url']=sheet.cell(i,2).value
			sub_data['data']=sheet.cell(i,3).value
			sub_data['method']=sheet.cell(i,4).value
			test_data.append(sub_data)
		return test_data


