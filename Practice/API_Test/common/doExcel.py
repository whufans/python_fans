#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/18 12:37
# @Author  : fans
# @File    : doExcel.py
# @Software: PyCharm Community Edition
import openpyxl
from API_Test.common.request import Request
import json
class Case:
	def __init__(self):
		self.id=None
		self.title=None
		self.method=None
		self.url=None
		self.data=None
		self.expected=None
		self.actual=None
		self.result=None

class DoExcel:
	def __init__(self,file_name):
		self.workbook=openpyxl.load_workbook(file_name)

	def  read_excel(self,sheet_name):
		sheet=self.workbook[sheet_name]
		max_row=sheet.max_row
		cases=[]
		for r in range(2,max_row+1):
			case=Case()
			case.id = sheet.cell(row=r,column=1).value
			case.title = sheet.cell(row=r, column=2).value
			case.method = sheet.cell(row=r, column=3).value
			case.url = sheet.cell(row=r, column=4).value
			case.data = sheet.cell(row=r, column=5).value
			case.expected = sheet.cell(row=r,column=6).value
			cases.append(case)
		return cases

	def write_excel(self,sheet_name,row,column,data):
		sheet = self.workbook[sheet_name]
		sheet.cell(row=row,column=column).value=data
		openpyxl.Workbook.save(self.workbook)

	def get_sheet_names(self):
		return self.workbook.sheetnames

if __name__ == '__main__':
	excel=DoExcel('test.xlsx')
	sheet_names=excel.get_sheet_names()
	print("sheetname:",sheet_names)
	for sheet_name in sheet_names:
		print("当前sheet:", sheet_name)
		cases=excel.read_excel(sheet_name)
		for case in cases:
			# print(case.__dict__)
			res=Request('',case.url,eval(case.data))
			print(res.resp)







