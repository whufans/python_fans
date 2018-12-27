#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/13 9:08
# @Author  : fans
# @File    : tdh_request.py
# @Software: PyCharm Community Edition
import base64
import json,jsonpath
from openpyxl import load_workbook
#通达海接口响应数据解码和核对流程：1、结果base64解密为XML  2、XML转为json   3、json base64解密 4、数据写入到用户表和组织表Excel，进行数据核对。
results=[]
#结果base64解密为XML
def   base64_decode2XML(rawData):
	decoder =base64.b64decode(rawData.encode('utf-8'))
	XMLData=str(decoder,'utf-8')
	return XMLData
#json核心节点base64解密
def json_decode(decode_list):
	data=[]
	for item  in decode_list:
		for key,value in item.items():
			try:
				item[key]=base64.b64decode(value.encode("utf-8")).decode("utf-8")
			except Exception as e:
				print(e)
		data.append(item)
	return data
#将通达海返回的用户信息写入适配层数据库导出的用户表中，便于核对数据正确性。
def write_user_info(file_path,sheet_name,data):
	wb=load_workbook(file_path)
	sheet = wb[sheet_name]
	row=sheet.max_row+1
	for i in range(1,len(data)+1):
		data_item=data[i-1]
		sheet.cell(row+i,2).value=data_item['FJM']
		sheet.cell(row+i,3).value=data_item['YHDM']
		sheet.cell(row+i,4).value=data_item['RYBS']
		sheet.cell(row+i,5).value=data_item['YHBM']
		sheet.cell(row+i,6).value=data_item['XM']
		sheet.cell(row+i,7).value=data_item['MD5']
		sheet.cell(row+i,8).value = data_item['LXFS']
		sheet.cell(row+i,9).value=data_item['SJHM']
		sheet.cell(row+i, 10).value = data_item['DZ']
		sheet.cell(row+i, 11).value = data_item['DZYX']
		sheet.cell(row+i, 12).value = data_item['SFZHM']
		sheet.cell(row+i, 13).value = data_item['XB']
		sheet.cell(row + i, 14).value = data_item['CSRQ']
		sheet.cell(row + i, 15).value = data_item['MZ']
		sheet.cell(row + i, 16).value = data_item['WHCD']
		sheet.cell(row + i, 17).value = data_item['ZZMM']
		sheet.cell(row + i, 18).value = data_item['XZJB']
		sheet.cell(row + i, 19).value = data_item['FGDJ']
		sheet.cell(row + i, 20).value = data_item['ZW']
		sheet.cell(row + i, 21).value = data_item['FJDJ']
		sheet.cell(row + i, 22).value = data_item['ZSBZ']
		sheet.cell(row + i, 23).value = data_item['SFPSY']
	wb.save(file_path)
#将通达海返回的组织信息写入适配层数据库导出的组织表中，便于核对数据正确性。
def write_org_info(file_path,sheet_name,data):
	wb=load_workbook(file_path)
	sheet = wb[sheet_name]
	row=sheet.max_row+1
	for i in range(1, len(data) + 1):
		data_item=data[i-1]
		sheet.cell(row+i,2).value = data_item['FY']
		sheet.cell(row+i,3).value=data_item['BMDM']
		sheet.cell(row+i,4).value=data_item['JGBS']
		sheet.cell(row+i,5).value=data_item['MC']
		sheet.cell(row+i,6).value=data_item['FBM']
		sheet.cell(row+i,7).value=data_item['JC']
		sheet.cell(row+i,8).value=data_item['LXR']
		sheet.cell(row+i,9).value=data_item['LXDH']
		sheet.cell(row+i,10).value=data_item['DZ']
		sheet.cell(row+i,11).value=data_item['PCFT']
	wb.save(file_path)
#运行函数：dataType:(1)'org' 组织  、'user' 用户（2）rawDataFile  通达海接口返回原始数据存放文件的地址
#（3）file_path 、sheet_name 数据写入Excel的文件地址和表单名
def run(dataType,rawDataFile,file_path,sheet_name):
	with open(rawDataFile,'r+') as file:
		global results
		rawData=file.readlines()[0]        #获取通达海接口返回数据
		XMLData=base64_decode2XML(rawData) #1、结果base64解密为XML
		jsonObj = json.loads(XMLData)      #2、XML转为json
		if dataType.upper()=='ORG':
			results=jsonpath.jsonpath(jsonObj,"$..ZZJG[*]")   #获取组织机构部分数据
			print(results)
		elif dataType.upper()=='USER':
			results = jsonpath.jsonpath(jsonObj, "$..RY[*]")  # 获取用户部分数据
			print(results)
		data=json_decode(results)          #3、json节点元素值 base64解密
		write_org_info(file_path,sheet_name,data)#数据写入到Excel

if __name__ == '__main__':
	#执行运行函数，进行数据写入，根据实际需求填写入参
	#zzjgData中存放通达海获取组织机构信息接口的返回结果
	# userData中存放通达海获取用户信息接口的返回结果
	run('org','zzjgData','tbl_court_organization.xlsx','tbl_court_organization')
	run('user', 'userData', 'tbl_court_user.xlsx', 'tbl_court_user')

