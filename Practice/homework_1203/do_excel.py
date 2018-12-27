#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/4 8:51
# @Author  : fans
# @File    : do_excel.py
# @Software: PyCharm Community Edition
# 1:根据老师上课讲解的测试用例参数化，去写单元测试用例以及用到超继承
# 2：把所有的测试用例数据存到Excel中，然后写一个类do_excel，完成测试数据的读取
# 3：利用do_excel类读取到的测试数据 传给1203课堂上讲解的test_suite，完成测试用例的加载。
# 4：要求：登录3条用例 充值3条用例
# 5：请给用例加上断言 以及异常处理，断言可以用code 也可以用msg或者是用整段信息都OK，最后要有测试报告出具
from   openpyxl import  load_workbook
class Do_excel:
	def __init__(self,file_path,sheet_name):
		self.file_path=file_path
		self.sheet_name=sheet_name
	def do_excel(self):
		workbook=load_workbook(self.file_path)
		sheet=workbook[self.sheet_name]
		test_data=[]
		for item in range(2,sheet.max_row+1):
			sub_data={}
			sub_data['id']=sheet.cell(item,1).value
			sub_data['http_method'] = sheet.cell(item, 2).value
			sub_data['url'] = sheet.cell(item, 3).value
			sub_data['data'] = sheet.cell(item, 4).value
			sub_data['expected'] = sheet.cell(item, 5).value
			test_data.append(sub_data)
		return test_data



